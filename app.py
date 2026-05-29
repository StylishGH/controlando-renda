import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import json
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

SAVE_FILE = os.path.join(os.path.dirname(__file__), "financas_dados.json")

MESES = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
]

CORES = {
    "bg": "#F7F6F3",
    "card": "#FFFFFF",
    "border": "#E0DDD8",
    "verde": "#1D9E75",
    "azul": "#378ADD",
    "amarelo": "#BA7517",
    "roxo": "#7F77DD",
    "rosa": "#D4537E",
    "cinza": "#888780",
    "vermelho": "#E24B4A",
    "texto": "#2C2C2A",
    "texto2": "#5F5E5A",
    "header_bg": "#2C2C2A",
    "header_fg": "#F7F6F3",
}

BALDES = [
    ("buffer",       "Buffer de emergência (casa)",    CORES["verde"]),
    ("ferramentas",  "Reserva ferramentas (pai)",       CORES["azul"]),
    ("parcelas",     "Abater parcelas extra",           CORES["amarelo"]),
    ("carro",        "Carro / investimento futuro",     CORES["roxo"]),
    ("lazer",        "Lazer / qualidade de vida",       CORES["rosa"]),
]

class FinancasApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Finanças da Casa")
        self.geometry("780x820")
        self.configure(bg=CORES["bg"])
        self.resizable(True, True)
        self.minsize(700, 700)

        self.mes_var = tk.StringVar(value=MESES[datetime.now().month - 1])
        self.ano_var = tk.StringVar(value=str(datetime.now().year))

        self.sliders = {}
        self.slider_labels = {}
        self.alloc_sliders = {}
        self.alloc_labels = {}
        self.alloc_amt_labels = {}
        self.metric_labels = {}

        self._build_ui()
        self._load_state()
        self._update()

    # ─── UI BUILD ────────────────────────────────────────────────

    def _build_ui(self):
        self._build_header()
        canvas = tk.Canvas(self, bg=CORES["bg"], highlightthickness=0)
        scroll = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.inner = tk.Frame(canvas, bg=CORES["bg"])
        self.inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.inner, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True, padx=0, pady=0)
        scroll.pack(side="right", fill="y")
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))

        pad = {"padx": 18, "pady": 6, "fill": "x"}
        self._section("RENDA MENSAL", self.inner, **pad)
        self._slider_row("renda", "Entrada do mês (R$)", 2000, 7000, 4000, self.inner, **pad)
        self._divider(self.inner, **pad)

        self._section("GASTOS FIXOS DA CASA", self.inner, **pad)
        self._slider_row("mercado",    "Mercado / alimentação",          300, 1500, 800,  self.inner, **pad)
        self._slider_row("contas",     "Contas (luz, água, net, gás)",   150,  800, 350,  self.inner, **pad)
        self._slider_row("transporte", "Transporte",                       0,  600, 200,  self.inner, **pad)
        self._slider_row("outros",     "Outros fixos",                     0,  800, 300,  self.inner, **pad)
        self._divider(self.inner, **pad)

        self._section("PARCELAS EM ABERTO", self.inner, **pad)
        self._slider_row("parcelas", "Total de parcelas/mês", 0, 2000, 700, self.inner, **pad)
        self._divider(self.inner, **pad)

        self._section("RESUMO DO MÊS", self.inner, **pad)
        self._build_metrics(self.inner, **pad)
        self.alert_label = tk.Label(self.inner, text="", bg=CORES["bg"], font=("Arial", 11),
                                    wraplength=680, justify="left", pady=4)
        self.alert_label.pack(**pad)
        self._divider(self.inner, **pad)

        self._section("ALOCAÇÃO DA SOBRA (%)", self.inner, **pad)
        for key, nome, cor in BALDES:
            self._alloc_row(key, nome, cor, self.inner, **pad)
        self.alloc_warn = tk.Label(self.inner, text="⚠ Total ultrapassa 100% — ajuste os percentuais.",
                                   fg=CORES["vermelho"], bg=CORES["bg"], font=("Arial", 10))
        self.alloc_warn.pack(padx=18, pady=2, anchor="w")
        self.alloc_warn.pack_forget()
        self._divider(self.inner, **pad)

        self._section("BALDES — DESTINO DA SOBRA", self.inner, **pad)
        self._build_baldes(self.inner, padx=18, pady=6)
        self._divider(self.inner, **pad)

        self._build_footer(self.inner)

    def _build_header(self):
        hdr = tk.Frame(self, bg=CORES["header_bg"], pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Finanças da Casa", bg=CORES["header_bg"], fg=CORES["header_fg"],
                 font=("Arial", 16, "bold")).pack(side="left", padx=20)
        right = tk.Frame(hdr, bg=CORES["header_bg"])
        right.pack(side="right", padx=20)
        tk.Label(right, text="Mês:", bg=CORES["header_bg"], fg=CORES["header_fg"],
                 font=("Arial", 11)).pack(side="left", padx=(0, 4))
        mes_cb = ttk.Combobox(right, textvariable=self.mes_var, values=MESES, width=10, state="readonly")
        mes_cb.pack(side="left", padx=(0, 8))
        tk.Label(right, text="Ano:", bg=CORES["header_bg"], fg=CORES["header_fg"],
                 font=("Arial", 11)).pack(side="left", padx=(0, 4))
        tk.Entry(right, textvariable=self.ano_var, width=6, font=("Arial", 11)).pack(side="left")

    def _section(self, title, parent, **pack_kwargs):
        tk.Label(parent, text=title, bg=CORES["bg"], fg=CORES["texto2"],
                 font=("Arial", 9), anchor="w").pack(**pack_kwargs)

    def _divider(self, parent, **pack_kwargs):
        tk.Frame(parent, bg=CORES["border"], height=1).pack(**pack_kwargs)

    def _slider_row(self, key, label, mn, mx, default, parent, **pack_kwargs):
        row = tk.Frame(parent, bg=CORES["bg"])
        row.pack(**pack_kwargs)
        tk.Label(row, text=label, bg=CORES["bg"], fg=CORES["texto2"],
                 font=("Arial", 11), width=30, anchor="w").pack(side="left")
        var = tk.IntVar(value=default)
        sl = tk.Scale(row, from_=mn, to=mx, resolution=50, orient="horizontal",
                      variable=var, bg=CORES["bg"], highlightthickness=0,
                      troughcolor=CORES["border"], activebackground=CORES["azul"],
                      showvalue=False, length=340, command=lambda v: self._update())
        sl.pack(side="left", padx=8)
        lbl = tk.Label(row, text=self._fmt(default), bg=CORES["bg"], fg=CORES["texto"],
                       font=("Arial", 11, "bold"), width=10, anchor="e")
        lbl.pack(side="left")
        self.sliders[key] = var
        self.slider_labels[key] = lbl

    def _alloc_row(self, key, label, cor, parent, **pack_kwargs):
        row = tk.Frame(parent, bg=CORES["bg"])
        row.pack(**pack_kwargs)
        dot = tk.Frame(row, bg=cor, width=10, height=10)
        dot.pack(side="left", padx=(0, 6))
        tk.Label(row, text=label, bg=CORES["bg"], fg=CORES["texto2"],
                 font=("Arial", 11), width=28, anchor="w").pack(side="left")
        var = tk.IntVar(value={"buffer":30,"ferramentas":20,"parcelas":20,"carro":15,"lazer":10}.get(key, 10))
        sl = tk.Scale(row, from_=0, to=100, resolution=5, orient="horizontal",
                      variable=var, bg=CORES["bg"], highlightthickness=0,
                      troughcolor=CORES["border"], activebackground=cor,
                      showvalue=False, length=200, command=lambda v: self._update())
        sl.pack(side="left", padx=8)
        pct_lbl = tk.Label(row, text="30%", bg=CORES["bg"], fg=CORES["texto"],
                           font=("Arial", 11, "bold"), width=4)
        pct_lbl.pack(side="left")
        amt_lbl = tk.Label(row, text="R$ 0", bg=CORES["bg"], fg=CORES["texto2"],
                           font=("Arial", 11), width=10, anchor="e")
        amt_lbl.pack(side="left")
        self.alloc_sliders[key] = var
        self.alloc_labels[key] = pct_lbl
        self.alloc_amt_labels[key] = amt_lbl

    def _build_metrics(self, parent, **pack_kwargs):
        frame = tk.Frame(parent, bg=CORES["bg"])
        frame.pack(**pack_kwargs)
        keys = [("fixos","Gastos fixos"), ("parc","Parcelas"), ("total","Total saída"), ("sobra","Sobra"), ("pct","Comprometido")]
        for i, (k, label) in enumerate(keys):
            card = tk.Frame(frame, bg=CORES["card"], bd=1, relief="flat",
                            highlightbackground=CORES["border"], highlightthickness=1)
            card.grid(row=0, column=i, padx=5, pady=4, sticky="nsew")
            frame.columnconfigure(i, weight=1)
            tk.Label(card, text=label, bg=CORES["card"], fg=CORES["texto2"],
                     font=("Arial", 9), pady=4).pack()
            lbl = tk.Label(card, text="—", bg=CORES["card"], fg=CORES["texto"],
                           font=("Arial", 13, "bold"), pady=4)
            lbl.pack()
            self.metric_labels[k] = lbl

    def _build_baldes(self, parent, **pack_kwargs):
        self.balde_frame = tk.Frame(parent, bg=CORES["bg"])
        self.balde_frame.pack(**pack_kwargs)
        self.balde_cards = {}
        all_baldes = BALDES + [("guardado", "Guardado (sem destino)", CORES["cinza"])]
        for i, (key, nome, cor) in enumerate(all_baldes):
            card = tk.Frame(self.balde_frame, bg=CORES["card"], bd=0,
                            highlightbackground=cor, highlightthickness=2)
            card.grid(row=i//3, column=i%3, padx=6, pady=6, sticky="nsew")
            self.balde_frame.columnconfigure(i%3, weight=1)
            tk.Frame(card, bg=cor, height=4).pack(fill="x")
            tk.Label(card, text=nome, bg=CORES["card"], fg=CORES["texto2"],
                     font=("Arial", 9), wraplength=160, justify="left", pady=4, padx=8).pack(anchor="w")
            val_lbl = tk.Label(card, text="R$ 0", bg=CORES["card"], fg=cor,
                               font=("Arial", 14, "bold"), padx=8)
            val_lbl.pack(anchor="w")
            pct_lbl = tk.Label(card, text="0% da sobra", bg=CORES["card"], fg=CORES["texto2"],
                               font=("Arial", 9), padx=8, pady=4)
            pct_lbl.pack(anchor="w")
            self.balde_cards[key] = (val_lbl, pct_lbl)

    def _build_footer(self, parent):
        footer = tk.Frame(parent, bg=CORES["bg"], pady=16)
        footer.pack(fill="x", padx=18)
        tk.Label(footer, text="Observações do mês:", bg=CORES["bg"], fg=CORES["texto2"],
                 font=("Arial", 10)).pack(anchor="w")
        self.obs_text = tk.Text(footer, height=3, font=("Arial", 11), bg=CORES["card"],
                                fg=CORES["texto"], relief="flat", bd=1,
                                highlightbackground=CORES["border"], highlightthickness=1)
        self.obs_text.pack(fill="x", pady=6)

        btn_frame = tk.Frame(footer, bg=CORES["bg"])
        btn_frame.pack(fill="x")
        tk.Button(btn_frame, text="💾  Salvar estado atual", command=self._save_state,
                  bg=CORES["azul"], fg="white", font=("Arial", 11, "bold"),
                  relief="flat", padx=16, pady=8, cursor="hand2").pack(side="left", padx=(0, 10))
        tk.Button(btn_frame, text="📊  Exportar planilha do mês", command=self._export_excel,
                  bg=CORES["verde"], fg="white", font=("Arial", 11, "bold"),
                  relief="flat", padx=16, pady=8, cursor="hand2").pack(side="left")
        tk.Button(btn_frame, text="🔄  Limpar mês", command=self._reset_month,
                  bg=CORES["bg"], fg=CORES["texto2"], font=("Arial", 10),
                  relief="flat", padx=12, pady=8, cursor="hand2").pack(side="right")

    # ─── LOGIC ────────────────────────────────────────────────────

    def _fmt(self, v):
        return f"R$ {int(v):,.0f}".replace(",", ".")

    def _update(self):
        renda = self.sliders["renda"].get()
        fixos = (self.sliders["mercado"].get() + self.sliders["contas"].get() +
                 self.sliders["transporte"].get() + self.sliders["outros"].get())
        parcelas = self.sliders["parcelas"].get()
        total = fixos + parcelas
        sobra = renda - total
        pct = round((total / renda) * 100) if renda else 0

        for key, var in self.sliders.items():
            self.slider_labels[key].config(text=self._fmt(var.get()))

        self.metric_labels["fixos"].config(text=self._fmt(fixos))
        self.metric_labels["parc"].config(text=self._fmt(parcelas))
        self.metric_labels["total"].config(text=self._fmt(total),
            fg=CORES["vermelho"] if pct > 90 else CORES["amarelo"] if pct > 75 else CORES["texto"])
        self.metric_labels["sobra"].config(text=self._fmt(sobra),
            fg=CORES["vermelho"] if sobra < 0 else CORES["amarelo"] if sobra < 400 else CORES["verde"])
        self.metric_labels["pct"].config(text=f"{pct}%",
            fg=CORES["vermelho"] if pct > 90 else CORES["amarelo"] if pct > 75 else CORES["texto"])

        if sobra < 0:
            self.alert_label.config(
                text=f"⚠ As saídas superam a renda em {self._fmt(abs(sobra))}. Revise os gastos antes de qualquer alocação.",
                fg=CORES["vermelho"])
        elif sobra < 400:
            self.alert_label.config(
                text="⚠ Margem muito baixa. Em meses fracos de serviço, vocês entram no negativo.",
                fg=CORES["amarelo"])
        else:
            self.alert_label.config(
                text="✓ Boa margem disponível. Distribua a sobra nos baldes abaixo.",
                fg=CORES["verde"])

        total_pct = sum(v.get() for v in self.alloc_sliders.values())
        if total_pct > 100:
            self.alloc_warn.pack(padx=18, pady=2, anchor="w")
        else:
            self.alloc_warn.pack_forget()

        sobra_pos = max(sobra, 0)
        used_pct = 0
        for key, nome, cor in BALDES:
            pct_balde = self.alloc_sliders[key].get()
            val = sobra_pos * pct_balde / 100
            self.alloc_labels[key].config(text=f"{pct_balde}%")
            self.alloc_amt_labels[key].config(text=self._fmt(val))
            val_lbl, pct_lbl = self.balde_cards[key]
            val_lbl.config(text=self._fmt(val))
            pct_lbl.config(text=f"{pct_balde}% da sobra")
            used_pct += pct_balde

        rest_pct = max(100 - used_pct, 0)
        rest_val = sobra_pos * rest_pct / 100
        val_lbl, pct_lbl = self.balde_cards["guardado"]
        val_lbl.config(text=self._fmt(rest_val))
        pct_lbl.config(text=f"{rest_pct}% da sobra")

    # ─── SAVE/LOAD ────────────────────────────────────────────────

    def _get_state(self):
        return {
            "mes": self.mes_var.get(),
            "ano": self.ano_var.get(),
            "sliders": {k: v.get() for k, v in self.sliders.items()},
            "alloc": {k: v.get() for k, v in self.alloc_sliders.items()},
            "obs": self.obs_text.get("1.0", "end").strip(),
        }

    def _save_state(self):
        with open(SAVE_FILE, "w", encoding="utf-8") as f:
            json.dump(self._get_state(), f, ensure_ascii=False, indent=2)
        messagebox.showinfo("Salvo", "Estado salvo com sucesso!")

    def _load_state(self):
        if not os.path.exists(SAVE_FILE):
            return
        try:
            with open(SAVE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            self.mes_var.set(data.get("mes", self.mes_var.get()))
            self.ano_var.set(data.get("ano", self.ano_var.get()))
            for k, v in data.get("sliders", {}).items():
                if k in self.sliders:
                    self.sliders[k].set(v)
            for k, v in data.get("alloc", {}).items():
                if k in self.alloc_sliders:
                    self.alloc_sliders[k].set(v)
            obs = data.get("obs", "")
            if obs:
                self.obs_text.delete("1.0", "end")
                self.obs_text.insert("1.0", obs)
        except Exception:
            pass

    def _reset_month(self):
        if messagebox.askyesno("Limpar mês", "Zerar todos os valores para um novo mês?"):
            defaults = {"renda":4000,"mercado":800,"contas":350,"transporte":200,"outros":300,"parcelas":700}
            for k, v in defaults.items():
                self.sliders[k].set(v)
            alloc_defaults = {"buffer":30,"ferramentas":20,"parcelas":20,"carro":15,"lazer":10}
            for k, v in alloc_defaults.items():
                self.alloc_sliders[k].set(v)
            self.obs_text.delete("1.0", "end")
            self._update()

    # ─── EXPORT ───────────────────────────────────────────────────

    def _export_excel(self):
        if not EXCEL_OK:
            messagebox.showerror("Erro", "Instale openpyxl:\n  pip install openpyxl")
            return

        state = self._get_state()
        renda = state["sliders"]["renda"]
        fixos = (state["sliders"]["mercado"] + state["sliders"]["contas"] +
                 state["sliders"]["transporte"] + state["sliders"]["outros"])
        parcelas_val = state["sliders"]["parcelas"]
        total = fixos + parcelas_val
        sobra = renda - total
        pct_comp = round((total / renda) * 100) if renda else 0
        sobra_pos = max(sobra, 0)

        mes_ano = f"{state['mes']}_{state['ano']}"
        default_name = f"financas_{mes_ano}.xlsx"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=default_name,
            title="Salvar planilha do mês"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = mes_ano

        # Styles
        def bold(size=11, color="2C2C2A"):
            return Font(name="Arial", bold=True, size=size, color=color)
        def reg(size=11, color="2C2C2A"):
            return Font(name="Arial", size=size, color=color)
        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color.replace("#",""))
        def border_bottom():
            return Border(bottom=Side(style="thin", color="E0DDD8"))
        center = Alignment(horizontal="center", vertical="center")
        left   = Alignment(horizontal="left",   vertical="center")
        right  = Alignment(horizontal="right",  vertical="center")
        brl = '#,##0.00" R$"'

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18

        row = 1
        ws.merge_cells(f"A{row}:C{row}")
        ws[f"A{row}"] = f"Finanças da Casa — {state['mes']} / {state['ano']}"
        ws[f"A{row}"].font = bold(14, "FFFFFF")
        ws[f"A{row}"].fill = fill("#2C2C2A")
        ws[f"A{row}"].alignment = center
        ws.row_dimensions[row].height = 30
        row += 1

        def section(label, color="378ADD"):
            nonlocal row
            row += 1
            ws.merge_cells(f"A{row}:C{row}")
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = bold(10, "FFFFFF")
            ws[f"A{row}"].fill = fill(color)
            ws[f"A{row}"].alignment = left
            ws.row_dimensions[row].height = 20
            row += 1

        def data_row(label, value, indent=False, formula=None, bold_row=False):
            nonlocal row
            ws[f"A{row}"] = ("  " if indent else "") + label
            ws[f"A{row}"].font = bold(11) if bold_row else reg(11)
            ws[f"A{row}"].alignment = left
            ws[f"B{row}"] = formula if formula else value
            ws[f"B{row}"].font = bold(11) if bold_row else reg(11)
            ws[f"B{row}"].alignment = right
            ws[f"B{row}"].number_format = brl
            ws[f"B{row}"].border = border_bottom()
            ws[f"A{row}"].border = border_bottom()
            ws.row_dimensions[row].height = 22
            row += 1

        section("RENDA MENSAL", "1D9E75")
        renda_row = row
        data_row("Entrada do mês", renda)

        section("GASTOS FIXOS", "378ADD")
        data_row("Mercado / alimentação", state["sliders"]["mercado"], indent=True)
        data_row("Contas (luz, água, net, gás)", state["sliders"]["contas"], indent=True)
        data_row("Transporte", state["sliders"]["transporte"], indent=True)
        data_row("Outros fixos", state["sliders"]["outros"], indent=True)
        first_fixo = renda_row + 2
        last_fixo = row - 1
        data_row("Subtotal fixos", fixos, formula=f"=SUM(B{first_fixo}:B{last_fixo})", bold_row=True)
        fixos_row = row - 1

        section("PARCELAS EM ABERTO", "BA7517")
        data_row("Total parcelas/mês", parcelas_val)
        parcelas_row = row - 1

        section("RESUMO", "7F77DD")
        data_row("Total saída", total, formula=f"=B{fixos_row}+B{parcelas_row}", bold_row=True)
        total_row = row - 1
        data_row("Sobra disponível", sobra, formula=f"=B{renda_row}-B{total_row}", bold_row=True)
        sobra_row = row - 1
        ws[f"B{sobra_row}"].font = Font(name="Arial", bold=True, size=11,
            color="1D9E75" if sobra >= 400 else "BA7517" if sobra >= 0 else "E24B4A")

        ws[f"A{row}"] = "% da renda comprometida"
        ws[f"A{row}"].font = reg(11)
        ws[f"B{row}"] = f"=B{total_row}/B{renda_row}"
        ws[f"B{row}"].number_format = "0.0%"
        ws[f"B{row}"].alignment = right
        ws[f"B{row}"].font = Font(name="Arial", size=11,
            color="E24B4A" if pct_comp > 90 else "BA7517" if pct_comp > 75 else "2C2C2A")
        row += 1

        section("ALOCAÇÃO DA SOBRA", "D4537E")
        ws[f"A{row}"] = "Destino"
        ws[f"B{row}"] = "% da sobra"
        ws[f"C{row}"] = "Valor (R$)"
        for cell in [ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"]]:
            cell.font = bold(10, "5F5E5A")
            cell.alignment = center
        row += 1

        used_pct = 0
        for key, nome, cor in BALDES:
            pct_b = state["alloc"][key]
            val_b = sobra_pos * pct_b / 100
            used_pct += pct_b
            ws[f"A{row}"] = nome
            ws[f"B{row}"] = pct_b / 100
            ws[f"B{row}"].number_format = "0%"
            ws[f"C{row}"] = f"=MAX(B{sobra_row},0)*B{row}"
            ws[f"C{row}"].number_format = brl
            for cell in [ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"]]:
                cell.font = reg(11)
                cell.alignment = right if cell.column > 1 else left
                cell.border = border_bottom()
            ws.row_dimensions[row].height = 22
            row += 1

        rest_pct = max(100 - used_pct, 0)
        ws[f"A{row}"] = "Guardado (sem destino)"
        ws[f"B{row}"] = rest_pct / 100
        ws[f"B{row}"].number_format = "0%"
        ws[f"C{row}"] = f"=MAX(B{sobra_row},0)*B{row}"
        ws[f"C{row}"].number_format = brl
        for cell in [ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"]]:
            cell.font = bold(11)
            cell.border = border_bottom()
            cell.alignment = right if cell.column > 1 else left
        ws.row_dimensions[row].height = 22
        row += 2

        obs = state["obs"]
        if obs:
            ws.merge_cells(f"A{row}:C{row}")
            ws[f"A{row}"] = "Observações:"
            ws[f"A{row}"].font = bold(10, "5F5E5A")
            row += 1
            ws.merge_cells(f"A{row}:C{row+2}")
            ws[f"A{row}"] = obs
            ws[f"A{row}"].font = reg(11)
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = 60
            row += 3

        ws.merge_cells(f"A{row}:C{row}")
        ws[f"A{row}"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
        ws[f"A{row}"].font = reg(9, "888780")
        ws[f"A{row}"].alignment = right

        wb.save(path)
        messagebox.showinfo("Exportado!", f"Planilha salva em:\n{path}")


if __name__ == "__main__":
    app = FinancasApp()
    app.mainloop()
